import sqlite3
import pandas as pd


class PeerEngine:

    def __init__(self, db_path="db/nifty100.db"):

        self.conn = sqlite3.connect(db_path)

    def load(self):

        ratios = pd.read_sql(
            "SELECT * FROM financial_ratios",
            self.conn
        )

        peers = pd.read_sql(
            "SELECT * FROM peer_groups",
            self.conn
        )

        return ratios, peers

    def compute(self):

        ratios, peers = self.load()

        df = ratios.merge(
            peers,
            on="company_id",
            how="left"
        )

        metrics = [

            "return_on_equity_pct",

            "return_on_capital_employed_pct",

            "net_profit_margin_pct",

            "debt_to_equity",

            "free_cash_flow_cr",

            "pat_cagr_5yr",

            "revenue_cagr_5yr",

            "eps_cagr_5yr",

            "interest_coverage",

            "asset_turnover"

        ]

        rows = []

        for metric in metrics:

            temp = df[
                [
                    "company_id",
                    "year",
                    "peer_group_name",
                    metric
                ]
            ].copy()

            if metric == "debt_to_equity":

                temp["percentile_rank"] = (

                    1 -

                    temp.groupby(
                        "peer_group_name"
                    )[metric]

                    .rank(pct=True)

                )

            else:

                temp["percentile_rank"] = (

                    temp.groupby(
                        "peer_group_name"
                    )[metric]

                    .rank(pct=True)

                )

            temp["metric"] = metric

            temp.rename(

                columns={
                    metric: "value"
                },

                inplace=True

            )

            rows.append(temp)

        out = pd.concat(
            rows,
            ignore_index=True
        )

        return out

    def save(self):

        df = self.compute()

        df.to_sql(

            "peer_percentiles",

            self.conn,

            if_exists="replace",

            index=False

        )

        return df

    def export_excel(self, df):

        import openpyxl
        from openpyxl.styles import PatternFill

        writer = pd.ExcelWriter(
            "output/peer_comparison.xlsx",
            engine="openpyxl"
        )

        green = PatternFill(
            fill_type="solid",
            fgColor="90EE90"
        )

        yellow = PatternFill(
            fill_type="solid",
            fgColor="FFFF99"
        )

        red = PatternFill(
            fill_type="solid",
            fgColor="FF9999"
        )

        peer_groups = sorted(
            df["peer_group_name"]
            .fillna("No Peer Group")
            .unique()
        )

        for group in peer_groups:

            sheet_df = df[
                df["peer_group_name"]
                .fillna("No Peer Group") == group
            ]

            sheet_df.to_excel(
                writer,
                sheet_name=str(group)[:31],
                index=False
            )

        writer.close()

        wb = openpyxl.load_workbook(
            "output/peer_comparison.xlsx"
        )

        for ws in wb.worksheets:

            headers = [
                cell.value
                for cell in ws[1]
            ]

            if "percentile_rank" not in headers:
                continue

            col = headers.index(
                "percentile_rank"
            ) + 1

            for row in range(2, ws.max_row + 1):

                cell = ws.cell(row, col)

                try:

                    value = float(cell.value)

                except:

                    continue

                if value >= 0.75:

                    cell.fill = green

                elif value <= 0.25:

                    cell.fill = red

                else:

                    cell.fill = yellow

        wb.save(
            "output/peer_comparison.xlsx"
        )


if __name__ == "__main__":

    engine = PeerEngine()

    data = engine.save()

    engine.export_excel(data)

    print("Peer comparison generated.")